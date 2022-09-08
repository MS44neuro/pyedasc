"""In this file, the analysis of 5 .abf files with 1 AP each, 
10 kHz sample rate is performed. All files come from the same 
waveform protocol and have an equal number of sweeps. 
All parameters are stored in an Excel file and the figures for 
general parameters and thresholds are saved."""
# Importing modules for file synchronization
import os
import sys
import glob
import time

# Importing modules for excel file writing
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
import string

# Creating paths to the module and .abf files
data_path = "/../data/abfs/drg_neurons/action_potentials/single_ap/"
pathToHere = os.path.abspath(os.path.dirname(__file__))
pathToData = os.path.abspath(pathToHere + data_path)
pathToModule = os.path.abspath(pathToHere + "/../../")

# Setting the path to the ap module
sys.path.insert(0, pathToModule)
from pyedasc import ap

# Set workbook
wb = Workbook()
ws = wb.active

# Set text file for time performance stats
f= open(pathToData + "/ap_test_time_performance.txt","w+")

# Configure columns and rows
rows = list(range(2, 100))
columns = string. ascii_uppercase

# First row        
ws.row_dimensions[1].height = 25
ws.freeze_panes = 'A2'

# Other rows
for row in rows:
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row].height = 47
    
# Columns    
for col in columns:   
    ws.column_dimensions[col].width = 47
    ws.column_dimensions[('A'+ col)].width = 47
fontObj1 = Font(name='Times New Roman', bold=True)

# Configure rows and columns beyond A-Z
row_1 = []   
row_2 = []         
for c in columns:
    row_1.append(c + '1')         
    row_2.append('A' + c + '1')
row_list = row_1 + row_2


totalAnalysisTime = []


# Looping through all .abf files sorted in the data path folder
for file in sorted(glob.glob(pathToData + "/*.abf")):

    par = ap.Parameters(abf_file=file, 
                        current_channel_number=0, 
                        voltage_channel_number=1,
                        absolute_rp_interval=[10, 500],
                        sweep_number=4)
    plot = ap.Plotting(abf_file=file, 
                       current_channel_number=0, 
                       voltage_channel_number=1,
                       absolute_rp_interval=[10, 500],
                       sweep_number=4) 
    
    # Load all parameters names and set column titles
    column_list = par.data_loader()[1]
    for l, c in zip(row_list, column_list):
        ws[l] = c
        ws[l].font = fontObj1
        
        
    # Timing starts here
    start_time = time.time()
        
    
    # Load all parameters
    rowData = par.data_loader()[0]
    ws.append(rowData)
    wb.save(pathToData + '/ap_parameters.xlsx')
    # Save the plots for each file
    plot.ap_plot(action='save', extension='.png')    
    plot.thresholds(action='save', extension='.png')
    
    
    #Timing ends here
    end_time = time.time()
    
    totalAnalysisTime.append(end_time - start_time)
    print("File", file, "analyzed in: ", round(end_time - start_time, 3), "seconds.")
    f.write(f"File {file} analyzed in: {round(end_time - start_time, 3)} seconds.\n")
    
    
print("Mean analysis time: ", round(np.mean(totalAnalysisTime), 3), " seconds.")
f.write(f"Mean analysis time: {round(np.mean(totalAnalysisTime), 3)} seconds.\n")

print("Total analysis time: ", round(sum(totalAnalysisTime), 3), "seconds.")
f.write(f"Total analysis time: {round(sum(totalAnalysisTime), 3)} seconds.\n")
    
f.close()  
  
 
    
 
    
